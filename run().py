import os
import subprocess
from openpyxl import load_workbook

unc_path = r'\\192.168.1.99\Tools1'
username = '/user:can'
password = 'freedom'

# Disconnect any existing connections to avoid error 1219
run(
    ['net', 'use', unc_path, '/delete'], 
    shell=True, 
    stdout=subprocess.DEVNULL, 
    stderr=subprocess.DEVNULL
)

# Connect to the UNC path with credentials
try:
    run(
        ['net', 'use', unc_path, password, username],
        shell=True,
        check=True
    )
except subprocess.CalledProcessError as e:
    print(f"Failed to connect to shared folder: {e}")
    exit(1)

# Build the file path and try to open the workbook
file_path = os.path.join(unc_path, 'test.xlsx')
print(f"Loading: {file_path}")

try:
    wb = load_workbook(file_path)
    ws = wb.active
    print("Cell A1:", ws['A1'].value)
except Exception as e:
    print(f"Failed to open workbook: {e}")
